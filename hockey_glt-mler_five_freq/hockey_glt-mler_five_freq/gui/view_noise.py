'''
Created on 13. Nov. 2016

@author: amueller
'''

from tkinter import N, S, W, E, IntVar, StringVar
from tkinter.ttk import Frame, Label, LabelFrame, OptionMenu, Checkbutton, Button,\
    Treeview, Entry
import numpy as np
from openpyxl import Workbook, load_workbook
from queue import Queue
from os.path import isfile

from goalref.reader_interface import NUM_FREQUENCIES
from goalref.moving_average import MovingAverage

class ViewNoise(Frame):
    '''
    classdocs
    '''
    
    OUTPUT_FILE = 'noise.xlsx'

    def __init__(self, master, application, reader, calibration, noiseWindowLen):
        '''
        Constructor
        '''
        Frame.__init__(self, master)
        self._application = application
        self._reader = reader
        self._calibration = calibration
        
        self.calibModes = ['None', 'Main', 'Frame']
        
        self.frequencies = []
        for i in range(NUM_FREQUENCIES):
            self.frequencies.append("%d - %.1f kHz" % (i, reader.getFrequency(i)/1000.0))
        
        self._averages = []
        for i in range(reader.getNumAntennas()):
            self._averages.append([MovingAverage(noiseWindowLen), MovingAverage(noiseWindowLen)])
        self._movingAverageUpdateCnt = 0
        
        self._calOffset = True
        self._calReference = True
        self._calCurrent = False
        self._freq = 0
        
        self._createWidgets()
        self._visible = True
        
        self._queue = Queue()
        self._dataRequest = reader.requestData(self._processSamples, blockSize=100, blocks=-1)
        self._checkQueue()
        
    def activate(self, active):
        self._visible = active
        
    def _createWidgets(self):
        self.rowconfigure(3, weight=1)
        self.columnconfigure(0, weight=1)
        
        frequencyFrame = LabelFrame(self, text="Frequency selection")
        frequencyFrame.grid(row=0, column=0, sticky=W+E, padx=10, pady=(10,0))
        frequencyFrame.columnconfigure(2, weight=1)
        
        self._varFreq = StringVar(self, value=self.frequencies[1])
        self._lblFreqTitle = Label(frequencyFrame, text="Frequency:")
        self._lblFreqTitle.grid(row=0, column=0, sticky=W+E, padx=(10, 5), pady=5)
        self._menuFreq = OptionMenu(frequencyFrame, self._varFreq, self.frequencies[1], *self.frequencies)
        self._menuFreq.grid(row=0, column=1, sticky=W+E, pady=5)
        
        self._btnScale = Button(frequencyFrame, text="Export noise figures", command=self._doExport)
        self._btnScale.grid(row=0, column=2, sticky=E, padx=10, pady=5)
        
        calibrationFrame = LabelFrame(self, text="Calibration control")
        calibrationFrame.grid(row=1, column=0, sticky=W+E, padx=10, pady=(5,0))
        calibrationFrame.columnconfigure(0, weight=1)
        calibrationFrame.columnconfigure(1, weight=1)
        
        self._varCalOffset = IntVar(self, value=1)
        self._checkCalOffset = Checkbutton(calibrationFrame, variable=self._varCalOffset, text="Apply offset cal.")
        self._checkCalOffset.grid(row=0, column=0, sticky=W+E, padx=10, pady=5)
        self._varCalReference = IntVar(self, value=1)
        self._checkCalReference = Checkbutton(calibrationFrame, variable=self._varCalReference, text="Apply reference cal.")
        self._checkCalReference.grid(row=0, column=1, sticky=W+E, padx=10, pady=5)
        self._varCalCurrent = IntVar(self, value=0)
        self._checkCalCurrent = Checkbutton(calibrationFrame, variable=self._varCalCurrent, text="Apply noise cancellation")
        self._checkCalCurrent.grid(row=0, column=2, sticky=W+E, padx=10, pady=5)
        self._btnRecalibrateOffset = Button(calibrationFrame, text="Recalibrate offsets", command=self._application.doOffsetRecalibration)
        self._btnRecalibrateOffset.grid(row=1, column=0, columnspan=3, sticky=W+E, padx=10, pady=(0,5))
        
        antennaCtrlFrame = LabelFrame(self, text="Antenna control")
        antennaCtrlFrame.grid(row=2, column=0, sticky=W+E, padx=10, pady=5)
        antennaCtrlFrame.columnconfigure(1, weight=1)
        
        self._varGain = StringVar(self, value="32")
        self._lblGain = Label(antennaCtrlFrame, text="Antenna gain:")
        self._lblGain.grid(row=0, column=0, sticky=W+E, padx=10, pady=5)
        self._txtGain = Entry(antennaCtrlFrame, width=10, textvariable=self._varGain)
        self._txtGain.grid(row=0, column=1, sticky=W+E, padx=10, pady=5)
        self._btnSetGain = Button(antennaCtrlFrame, text="Set gain", command=self._setGain)
        self._btnSetGain.grid(row=0, column=2, sticky=W+E, padx=10, pady=5)
        
        self._varCalib = StringVar(self, value=self.calibModes[0])
        self._lblCalib = Label(antennaCtrlFrame, text="Cal. loop selection:")
        self._lblCalib.grid(row=1, column=0, sticky=W+E, padx=(10,5), pady=5)
        self._menuCalib = OptionMenu(antennaCtrlFrame, self._varCalib, self.calibModes[0], *self.calibModes)
        self._menuCalib.grid(row=1, column=1, sticky=W+E, padx=10, pady=5)
        self._btnSetGain = Button(antennaCtrlFrame, text="Set cal. loops", command=self._setCalib)
        self._btnSetGain.grid(row=1, column=2, sticky=W+E, padx=10, pady=5)
        
        noiseFrame = LabelFrame(self, text="Noise levels")
        noiseFrame.grid(row=3, column=0, sticky=N+S+W+E, padx=10, pady=(5,10))
        noiseFrame.rowconfigure(0, weight=1)
        noiseFrame.columnconfigure(0, weight=1)
        
        self._cols = ["Main mean ampl.", "Main mean phase", "Main total SD", "Main variation",
            "Frame mean ampl.", "Frame mean phase", "Frame total SD", "Frame variation"]
        self._noiseTree = Treeview(noiseFrame, columns=self._cols,
                                   displaycolumns="#all", height=self._reader.getNumAntennas())
        self._noiseTree.grid(row=0, column=0, sticky=W+E+S+N, padx=10, pady=10)
        self._noiseTree.column("#0", stretch=False, minwidth=0, width=100)
        for key, val in enumerate(self._cols):
            self._noiseTree.column(key, stretch=True, minwidth=0, width=1)
            self._noiseTree.heading(key, text=val, anchor=W)
        for i in range(self._reader.getNumAntennas()):
            self._noiseTree.insert("", "end", iid='%d' % i, text='Antenna %s' % chr(ord('A') + i))
        
    def _checkQueue(self):
        # Retrieve data from queue and forward to MovingAverages for processing
        while not self._queue.empty():
            data = self._queue.get()
            for i, a in enumerate(self._averages):
                # Frame antenna is in even channels
                a[1].processSamples([sample.getSampleVal(i*2, self._freq) for sample in data])
                # Main antenna is in odd channels
                a[0].processSamples([sample.getSampleVal(i*2+1, self._freq) for sample in data])
        
        # Update table with figures from MovingAverages
        self._movingAverageUpdateCnt += 1
        if self._movingAverageUpdateCnt >= 25 and self._visible:
            self._output_vals = []
            for i, a in enumerate(self._averages):
                a0 = a[0].getCurrentAverage()
                a1 = a[1].getCurrentAverage()
                np0 = a[0].getNoisePower()
                np1 = a[1].getNoisePower()
                vals = [
                        np.abs(a0),
                        np.rad2deg(np.angle(a0)),
                        np.sqrt(np0),
                        np.sqrt(np0) / np.abs(a0),
                        np.abs(a1),
                        np.rad2deg(np.angle(a1)),
                        np.sqrt(np1),
                        np.sqrt(np1) / np.abs(a1)
                    ]
                self._noiseTree.item(i, values=[
                        '%.5e' % vals[0],
                        '%.2f' % vals[1],
                        '%.5e' % vals[2],
                        '%.5e' % vals[3],
                        '%.5e' % vals[4],
                        '%.2f' % vals[5],
                        '%.5e' % vals[6],
                        '%.5e' % vals[7]
                    ])
                self._output_vals.extend(vals)
            self._movingAverageUpdateCnt = 0
        
        # Update configuration data from GUI elements
        before = [self._calOffset, self._calReference, self._calCurrent, self._freq]
        
        self._calOffset = self._varCalOffset.get() == 1
        self._calReference = self._varCalReference.get() == 1
        self._calCurrent = self._varCalCurrent.get() == 1
        self._freq = self.frequencies.index(self._varFreq.get())
            
        after = [self._calOffset, self._calReference, self._calCurrent, self._freq]
        if before != after:
            while not self._queue.empty():
                self._queue.get()
            for a in self._averages:
                a[0].clear()
                a[1].clear()
        
        # Reschedule method execution
        self.after(25, self._checkQueue)

    def _processSamples(self, samples):
        out = []
        for sample in samples:
            self._calibration.apply(sample, offset=self._calOffset, reference=self._calReference, current=self._calCurrent)
            out.append(sample)
        
        self._queue.put(np.array(out))
        
    def _doExport(self):
        if not isfile(self.OUTPUT_FILE):
            wb = Workbook(self.OUTPUT_FILE)
            ws = wb.create_sheet('Noise figures')
            
            headline = ['Frequency [Hz]', 'Current [A]']
            for i in range(self._reader.getNumAntennas()):
                ant = chr(ord('A') + i)
                headline.extend(['Ant. %s %s' % (ant, x) for x in self._cols])
            ws.append(headline)
        else:
            wb = load_workbook(self.OUTPUT_FILE)
            ws = wb['Noise figures']
            
        row = [self._reader.getFrequency(self._freq), self._reader.getExciterCurrent(self._freq)]
        row.extend(self._output_vals)
        
        ws.append(row)
        wb.save(self.OUTPUT_FILE)
        
    def _setGain(self):
        try:
            gain = int(self._varGain.get())
            for i in range(12):
                self._reader.setMainGain(i, gain)
                self._reader.setFrameGain(i, gain)
        except:
            pass
        
    def _setCalib(self):
        mode = self._varCalib.get()
        if mode == 'Main':
            for i in range(12):
                self._reader.setMainCalib(i, 1)
        elif mode == 'Frame':
            for i in range(12):
                self._reader.setFrameCalib(i, 1)
        else:
            for i in range(12):
                self._reader.setMainCalib(i, 0)
        
